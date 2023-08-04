from xlrd import open_workbook
import re
from xlutils.copy import copy
from datetime import datetime
import xlwt
from openpyxl import load_workbook

def readExcel(filename, n):
    excel = open_workbook(filename)
    sheet = excel.sheets()[0]
    col = sheet.col_values(n)  # n=12:waring content ; n=44:state of kaidan
    return col


# excel未处理不需要的告警！
# 归纳合并所需的告警类信息
def extractData(col):
    bracket = re.compile('(\()')
    idPattern = re.compile(r'(\()(.*?)(\))')
    timePattern = re.compile('.* ([0-9][0-9]):([0-9][0-9]).*')
    MON_OBJ_ID = []
    timeHour = []
    p1Dict = {}
    # p2Dict = {}
    count = 1
    for p in col[1:]:
        errorList = []
        if "交易量下降" in p or "交易下降了" in p:
            errorList.append("fluctValue")
        if "秒无交易上送" in p:
            errorList.append("noTransTm")
        if "ERROR_CD" in p:
            errorList.append("httpCnt")
        if "秒内成功率" in p:
            errorList.append("succRate")
        if "秒内失败笔数" in p:
            errorList.append("transFailNum")
        if "连续失败笔数" in p:
            errorList.append("seqFailNum")
        try:
            objId = idPattern.search(p).group(2)
            while len(objId) < 5:
                ret = bracket.sub('', p, 1)
                objId = idPattern.search(ret).group(2)
                p = ret
            objTime = str(timePattern.search(p).group(1))
            # print(objTime)
            MON_OBJ_ID.append(objId)
            timeHour.append(objTime)
            # print(objTime,errorList,count)
        except:
            print("illegal data format.")
        if objId not in p1Dict:
            p1Dict[objId] = {}
        # print(errorList,objId)
        for ruleType in errorList:
            if ruleType not in p1Dict[objId]:
                p1Dict[objId][ruleType] = []
            if len(objTime) == 2 and objTime[0] == '0':
                objTime = objTime[1]
            p1Dict[objId][ruleType].append(objTime)
            # print(p2Dict,count,objId)
        # p1Dict[objId]=p2Dict
    print(p1Dict)
    return p1Dict


# 写入auto_rule文件
def writeExcelRule(p1Dict):
    nowTime = datetime.now().strftime("%m%d")
    fileName = "auto_rule_conf_" + nowTime + '.xls'
    workBook = xlwt.Workbook(encoding='utf-8')
    sheet = workBook.add_sheet('规则调优')
    sheet.write(0, 0, 'MON_OBJ_ID')
    sheet.write(0, 1, 'DATA_SOURCE_ID')
    sheet.write(0, 2, 'RULE_TYPE')
    sheet.write(0, 3, 'METHOD')
    sheet.write(0, 4, 'TIME')
    sheet.write(0, 5, 'MULTI_TRIG')
    sheet.write(0, 6, 'Frequency')
    workBook.save(fileName)
    autoRuleNum = 1
    for id in p1Dict:

        if id in ['UPACPNOTIFY_UPACPNOTIFY_NA', 'UPACP_UPACP_NA', 'UPACPTRANS_UPACPTRANS_NA']:
            pass
        elif 'UPACP' in id and 'Trans' in id:
            pass
        elif 'UPACP' in id and 'PS' in id:
            pass
        elif 'UPACP_ClientID' in id:
            pass
        elif 'UPACP_' in id and '_M' in id:
            pass
        elif id == 'HTTP_UPACP_SYSTEM':
            pass
        elif id in ['HTTP_upacqz_hlw', 'HTTP_upacqz_zx']:
            pass
        elif id == "QRC_QRC_NA":
            pass
        elif id in ['QRC_U_NA', 'QRC_H_NA']:
            pass
        elif 'QRC_QRCUP' in id or id == 'QRC_01_NA':
            pass
        elif 'QRC_' in id and '_U_NA' in id:
            pass
        elif 'QRC_' in id and '_H_NA' in id:
            pass
        elif id == 'HTTP_qrc_system':
            pass
        elif id in ['HTTP_qrc_zhuanxian_000001', 'HTTP_qrc_hlw_000001']:
            pass
        elif 'HTTP_qrc_system_ERROR_CD' in id:
            pass
        elif id == 'FS_FS_NA':
            pass
        elif 'FS_CENTER_Center' in id:
            pass
        elif 'FS_TYPE_' in id:
            pass
        elif 'FS_PROC_SYS_' in id:
            pass
        elif id == 'HTTP_fsas_system':
            pass
        elif id in ['HTTP_fsas_zx', 'HTTP_fsas_hlw']:
            pass
        elif 'HTTP_fsas_system_ERROR_CD' in id:
            pass
        elif id == 'QTT_QTTB_B':
            pass
        elif 'QTT_CENTER_S' in id:
            pass
        elif id in ['QTT_JSAPI_B', 'QTT_MICROPAY_B', 'QTT_NATIVE_B', 'QTT_APP_B']:
            pass
        elif id == 'HTTP_T_System':
            pass
        elif id == 'HTTP_QTT_DOMAINM_':
            pass
        elif id in ['HTTP_T_QZ_ZX', 'HTTP_T_QZ_HLW']:
            pass
        elif 'HTTP_T_QZ_ZX_' in id or 'HTTP_T_QZ_HLW_' in id:
            pass
        elif 'HTTP_TSystem_ERROR_CD_' in id:
            pass
        elif id == 'QAT_QAT_NA':
            pass
        elif id == 'QAT_CENTER_SH':
            pass
        elif id in ['QAT_pay_B', 'QAT_create_B', 'QAT_precreate_B']:
            pass
        elif id == 'HTTP_A_System':
            pass
        elif id in ['HTTP_A_QZ_ZX', 'HTTP_A_QZ_HLW']:
            pass
        elif id not in ['HTTP_A_QZ_ZX', 'HTTP_A_QZ_HLW'] and 'HTTP_A_QZ_' in id:
            pass
        elif 'HTTP_ASystem_ERROR_CD_' in id:
            pass
# -----------20190829update--------------------------------
        elif 'QRC_' in id and '_IM' in id:
            pass
        elif 'QAT_' in id and '_B' in id:
            pass
        elif 'QTT_' in id and '_B' in id:
            pass
        elif 'UPACP_' in id and '_R' in id:
            pass
        elif 'UPACP_' in id and '_A' in id:
            pass
        # ---------------------------upadate 20190829-----------------
# -----------20190902update--------------------------------
        elif 'FS_T_' in id and '_A' in id:
            continue
        elif 'FS_T_' in id and '_R' in id:
            continue
        elif 'FS_A_' in id and '_R' in id:
            continue
        elif 'FS_A_' in id and '_A' in id:
            continue
        elif 'FS_' in id and '_A' in id:
            pass
        elif 'FS_' in id and '_R' in id:
            pass
# -----------20190902update--------------------------------
# -----------20190917update--------------------------------
        elif 'UPACPTRANS_orderN_' in id or 'UPACPNOTIFY_notice_' in id:
            pass
# -----------20190917update--------------------------------
        else:
            continue
        pattern = re.compile('(.*?)_')
        try:
            sourceId = pattern.match(id).group(1)
        except:
            print('Data fail')
            continue
        for errorType in p1Dict[id]:
            oldbook = open_workbook(fileName)
            wb = copy(oldbook)
            ws = wb.get_sheet(0)
            ws.write(autoRuleNum, 0, id)
            ws.write(autoRuleNum, 1, sourceId)
            ws.write(autoRuleNum, 2, errorType)
            ws.write(autoRuleNum, 3, 'auto')
            # 写入符合规格的时间信息
            seq = len(p1Dict[id][errorType])
            set1 = set(p1Dict[id][errorType])
            setlist = list(set1)
            setlist.sort()
            timeLog = ",".join(setlist)
            ws.write(autoRuleNum, 4, timeLog)
            ws.write(autoRuleNum, 5, '1')
            ws.write(autoRuleNum, 6, str(seq))
            wb.save(fileName)
            autoRuleNum += 1

# 写入告警分析模板
def writeExcelAnalyse(col_content, col_event):
    UPACP_All_Warning = UPACP_All_Event = UPACP_All_object = 0
    UPACP_All_object_list = []
    UPACP_Type_Warning = UPACP_Type_Event = UPACP_Type_object = 0
    UPACP_Type_object_list = []
    UPACP_procsys_Warning = UPACP_procsys_Event = UPACP_procsys_object = 0
    UPACP_procsys_object_list = []
    UPACP_client_id_Warning = UPACP_client_id_Event = UPACP_client_id_object = 0
    UPACP_client_id_object_list = []
    UPACP_Merchant_Warning = UPACP_Merchant_Event = UPACP_Merchant_object = 0
    UPACP_Merchant_object_list = []
    HTTP_UPACP_Warning = HTTP_UPACP_Event = HTTP_UPACP_object = 0
    HTTP_UPACP_object_list = []
    HTTP_UPACP_qzhlw_Warning = HTTP_UPACP_qzhlw_Event = HTTP_UPACP_qzhlw_object = 0
    HTTP_UPACP_qzhlw_object_list = []

    QRC_QRC_Warning = QRC_QRC_Event = QRC_QRC_object = 0
    QRC_QRC_object_list = []
    QRC_Center_Warning = QRC_Center_Event = QRC_Center_object = 0
    QRC_Center_object_list = []
    QRC_Type_Warning = QRC_Type_Event = QRC_Type_object = 0
    QRC_Type_object_list = []
    QRC_Order_Warning = QRC_Order_Event = QRC_Order_object = 0
    QRC_Order_object_list = []
    HTTP_QRC_All_Warning = HTTP_QRC_All_Event = HTTP_QRC_All_object = 0
    HTTP_QRC_All_object_list = []
    HTTP_QRC_zxhlw_Warning = HTTP_QRC_zxhlw_Event = HTTP_QRC_zxhlw_object = 0
    HTTP_QRC_zxhlw_object_list = []
    HTTP_QRC_error_Warning = HTTP_QRC_error_Event = HTTP_QRC_error_object = 0
    HTTP_QRC_error_object_list = []

    FS_FS_waring = FS_FS_Event = FS_FS_object = 0
    FS_FS_object_list = []
    FS_Center_waring = FS_Center_Event = FS_Center_object = 0
    FS_Center_object_list = []
    FS_Type_waring = FS_Type_Event = FS_Type_object = 0
    FS_Type_object_list = []
    FS_procSys_waring = FS_procSys_Event = FS_procSys_object = 0
    FS_procSys_object_list = []
    HTTP_fsasAll_waring = HTTP_fsasAll_Event = HTTP_fsasAll_object = 0
    HTTP_fsasAll_object_list = []
    HTTP_FS_qzhlw_waring = HTTP_FS_qzhlw_Event = HTTP_FS_qzhlw_object = 0
    HTTP_FS_qzhlw_object_list = []
    HTTP_FS_error_waring = HTTP_FS_error_Event = HTTP_FS_error_object = 0
    HTTP_FS_error_object_list = []

    QTT_QTTB_waring = QTT_QTTB_Event = QTT_QTTB_object = 0
    QTT_QTTB_object_list = []
    QTT_Center_waring = QTT_Center_Event = QTT_Center_object = 0
    QTT_Center_object_list = []
    QTT_Type_waring = QTT_Type_Event = QTT_Type_object = 0
    QTT_Type_object_list = []
    HTTP_QTTall_waring = HTTP_QTTall_Event = HTTP_QTTall_object = 0
    HTTP_QTTall_object_list = []
    HTTP_QTTdomain_waring = HTTP_QTTdomain_Event = HTTP_QTTdomain_object = 0
    HTTP_QTTdomain_object_list = []
    HTTP_QTThlwqz_waring = HTTP_QTThlwqz_Event = HTTP_QTThlwqz_object = 0
    HTTP_QTThlwqz_object_list = []
    HTTP_QTThlw_qz_waring = HTTP_QTThlw_qz_Event = HTTP_QTThlw_qz_object = 0
    HTTP_QTThlw_qz_object_list = []
    HTTP_QTT_error_waring = HTTP_QTT_error_Event = HTTP_QTT_error_object = 0
    HTTP_QTT_error_object_list = []

    QAT_QAT_waring = QAT_QAT_Event = QAT_QAT_object = 0
    QAT_QAT_object_list = []
    QAT_Center_waring = QAT_Center_Event = QAT_Center_object = 0
    QAT_Center_object_list = []
    QAT_Type_waring = QAT_Type_Event = QAT_Type_object = 0
    QAT_Type_object_list = []
    HTTP_QATall_waring = HTTP_QATall_Event = HTTP_QATall_object = 0
    HTTP_QATall_object_list = []
    HTTP_QATdomain_waring = HTTP_QATdomain_Event = HTTP_QATdomain_object = 0
    HTTP_QATdomain_object_list = []
    HTTP_QATAPI_waring = HTTP_QATAPI_Event = HTTP_QATAPI_object = 0
    HTTP_QATAPI_object_list = []
    HTTP_QAT_error_waring = HTTP_QAT_error_Event = HTTP_QAT_error_object = 0
    HTTP_QAT_error_object_list = []

#----------------------------------update 20190829----------------------------------
    UPACP_Institution_warning = UPACP_Institution_Event = UPACP_Institution_object = 0
    UPACP_Institution_list = []

    QRC_APP_warning = QRC_APP_Event = QRC_APP_object = 0
    QRC_APP_list = []

    QTT_Institution_warning = QTT_Institution_Event = QTT_Institution_object = 0
    QTT_Institution_list = []

    QAT_Institution_warning = QAT_Institution_Event = QAT_Institution_object = 0
    QAT_Institution_list = []
#----------------------------------update 20190829----------------------------------
# ----------------------------------update 20190829 新增客结机构130----------------------------------
    FS_Institution_warning = FS_Institution_Event = FS_Institution_object = 0
    FS_Institution_list = []
# ----------------------------------update 20190829----------------------------------
# ----------------------------------update 20190917----------------------------------
    UPACP_Center_warning = UPACP_Center_Event = UPACP_Center_object = 0
    UPACP_Center_list = []
# ----------------------------------update 20190917----------------------------------

    idPattern = re.compile(r'(\()(.*?)(\))')
    bracket = re.compile('(\()')
    index = 1

    # 按id归类，待补充
    for content in col_content[1:]:
        try:
            objId = idPattern.search(content).group(2)
            while len(objId) < 5:
                ret = bracket.sub('', content, 1)
                objId = idPattern.search(ret).group(2)
                content = ret
        except:
            print("illegal data format.")
            continue
        if objId in ['UPACPNOTIFY_UPACPNOTIFY_NA', 'UPACP_UPACP_NA', 'UPACPTRANS_UPACPTRANS_NA']:
            UPACP_All_Warning += 1
            if col_event[index] == '已开单':
                UPACP_All_Event += 1
            UPACP_All_object_list.append(objId)
        elif 'UPACP' in objId and 'Trans' in objId:
            UPACP_Type_Warning += 1
            if col_event[index] == '已开单':
                UPACP_Type_Event += 1
            UPACP_Type_object_list.append(objId)
        elif 'UPACP' in objId and 'PS' in objId:
            UPACP_procsys_Warning += 1
            if col_event[index] == '已开单':
                UPACP_procsys_Event += 1
            UPACP_procsys_object_list.append(objId)
        elif 'UPACP_ClientID' in objId:
            UPACP_client_id_Warning += 1
            if col_event[index] == '已开单':
                UPACP_client_id_Event += 1
            UPACP_client_id_object_list.append(objId)
        elif 'UPACP_' in objId and '_M' in objId:
            UPACP_Merchant_Warning += 1
            if col_event[index] == '已开单':
                UPACP_Merchant_Event += 1
            UPACP_Merchant_object_list.append(objId)
        elif objId == 'HTTP_UPACP_SYSTEM':
            HTTP_UPACP_Warning += 1
            if col_event[index] == '已开单':
                HTTP_UPACP_Event += 1
            HTTP_UPACP_object_list.append(objId)
        elif objId in ['HTTP_upacqz_hlw', 'HTTP_upacqz_zx']:
            HTTP_UPACP_qzhlw_Warning += 1
            if col_event[index] == '已开单':
                HTTP_UPACP_qzhlw_Event += 1
            HTTP_UPACP_qzhlw_object_list.append(objId)
        elif objId == "QRC_QRC_NA":
            QRC_QRC_Warning += 1
            if col_event[index] == '已开单':
                QRC_QRC_Event += 1
            QRC_QRC_object_list.append(objId)
        elif objId in ['QRC_U_NA', 'QRC_H_NA']:
            QRC_Center_Warning += 1
            if col_event[index] == '已开单':
                QRC_Center_Event += 1
            QRC_Center_object_list.append(objId)
        elif 'QRC_QRCUP' in objId or objId == 'QRC_01_NA':
            QRC_Type_Warning += 1
            if col_event[index] == '已开单':
                QRC_Type_Event += 1
            QRC_Type_object_list.append(objId)
        elif 'QRC_' in objId and '_U_NA' in objId:
            QRC_Order_Warning += 1
            if col_event[index] == '已开单':
                QRC_Order_Event += 1
            QRC_Order_object_list.append(objId)
        elif 'QRC_' in objId and '_H_NA' in objId:
            QRC_Order_Warning += 1
            if col_event[index] == '已开单':
                QRC_Order_Event += 1
            QRC_Order_object_list.append(objId)
        elif objId == 'HTTP_qrc_system':
            HTTP_QRC_All_Warning += 1
            if col_event[index] == '已开单':
                HTTP_QRC_All_Event += 1
            HTTP_QRC_All_object_list.append(objId)
        elif objId in ['HTTP_qrc_zhuanxian_000001', 'HTTP_qrc_hlw_000001']:
            HTTP_QRC_zxhlw_Warning += 1
            if col_event[index] == '已开单':
                HTTP_QRC_zxhlw_Event += 1
            HTTP_QRC_zxhlw_object_list.append(objId)
        elif 'HTTP_qrc_system_ERROR_CD' in objId:
            HTTP_QRC_error_Warning += 1
            if col_event[index] == '已开单':
                HTTP_QRC_error_Event += 1
            HTTP_QRC_error_object_list.append(objId)
        elif objId == 'FS_FS_NA':
            FS_FS_waring += 1
            if col_event[index] == '已开单':
                FS_FS_Event += 1
            FS_FS_object_list.append(objId)
        elif 'FS_CENTER_Center' in objId:
            FS_Center_waring += 1
            if col_event[index] == '已开单':
                FS_Center_Event += 1
            FS_Center_object_list.append(objId)
        elif 'FS_TYPE_' in objId:
            FS_Type_waring += 1
            if col_event[index] == '已开单':
                FS_Type_Event += 1
            FS_Type_object_list.append(objId)
        elif 'FS_PROC_SYS_' in objId:
            FS_procSys_waring += 1
            if col_event[index] == '已开单':
                FS_procSys_Event += 1
            FS_procSys_object_list.append(objId)
        elif objId == 'HTTP_fsas_system':
            HTTP_fsasAll_waring += 1
            if col_event[index] == '已开单':
                HTTP_fsasAll_Event += 1
            HTTP_fsasAll_object_list.append(objId)
        elif objId in ['HTTP_fsas_zx', 'HTTP_fsas_hlw']:
            HTTP_FS_qzhlw_waring += 1
            if col_event[index] == '已开单':
                HTTP_FS_qzhlw_Event += 1
            HTTP_FS_qzhlw_object_list.append(objId)
        elif 'HTTP_fsas_system_ERROR_CD' in objId:
            HTTP_FS_error_waring += 1
            if col_event[index] == '已开单':
                HTTP_FS_error_Event += 1
            HTTP_FS_error_object_list.append(objId)
        elif objId == 'QTT_QTTB_B':
            QTT_QTTB_waring += 1
            if col_event[index] == '已开单':
                QTT_QTTB_Event += 1
            QTT_QTTB_object_list.append(objId)
        elif 'QTT_CENTER_S' in objId:
            QTT_Center_waring += 1
            if col_event[index] == '已开单':
                QTT_Center_Event += 1
            QTT_Center_object_list.append(objId)
        elif objId in ['QTT_JSAPI_B', 'QTT_MICROPAY_B', 'QTT_NATIVE_B', 'QTT_APP_B']:
            QTT_Type_waring += 1
            if col_event[index] == '已开单':
                QTT_Type_Event += 1
            QTT_Type_object_list.append(objId)
        elif objId == 'HTTP_T_System':
            HTTP_QTTall_waring += 1
            if col_event[index] == '已开单':
                HTTP_QTTall_Event += 1
            HTTP_QTTall_object_list.append(objId)
        elif objId == 'HTTP_QTT_DOMAINM_':
            HTTP_QTTdomain_waring += 1
            if col_event[index] == '已开单':
                HTTP_QTTdomain_Event += 1
            HTTP_QTTdomain_object_list.append(objId)
        elif objId in ['HTTP_T_QZ_ZX', 'HTTP_T_QZ_HLW']:
            HTTP_QTThlwqz_waring += 1
            if col_event[index] == '已开单':
                HTTP_QTThlwqz_Event += 1
            HTTP_QTThlwqz_object_list.append(objId)
        elif 'HTTP_T_QZ_ZX_' in objId or 'HTTP_T_QZ_HLW_' in objId:
            HTTP_QTThlw_qz_waring += 1
            if col_event[index] == '已开单':
                HTTP_QTThlw_qz_Event += 1
            HTTP_QTThlw_qz_object_list.append(objId)
        elif 'HTTP_TSystem_ERROR_CD_' in objId:
            HTTP_QTT_error_waring += 1
            if col_event[index] == '已开单':
                HTTP_QTT_error_Event += 1
            HTTP_QTT_error_object_list.append(objId)
        elif objId == 'QAT_QAT_NA':
            QAT_QAT_waring += 1
            if col_event[index] == '已开单':
                QAT_QAT_Event += 1
            QAT_QAT_object_list.append(objId)
        elif objId == 'QAT_CENTER_SH':
            QAT_Center_waring += 1
            if col_event[index] == '已开单':
                QAT_Center_Event += 1
            QAT_Center_object_list.append(objId)
        elif objId in ['QAT_pay_B', 'QAT_create_B', 'QAT_precreate_B']:
            QAT_Type_waring += 1
            if col_event[index] == '已开单':
                QAT_Type_Event += 1
            QAT_Type_object_list.append(objId)
        elif objId == 'HTTP_A_System':
            HTTP_QATall_waring += 1
            if col_event[index] == '已开单':
                HTTP_QATall_Event += 1
            HTTP_QATall_object_list.append(objId)
        elif objId in ['HTTP_A_QZ_ZX', 'HTTP_A_QZ_HLW']:
            HTTP_QATdomain_waring += 1
            if col_event[index] == '已开单':
                HTTP_QATdomain_Event += 1
            HTTP_QATdomain_object_list.append(objId)
        elif objId not in ['HTTP_A_QZ_ZX', 'HTTP_A_QZ_HLW'] and 'HTTP_A_QZ_' in objId:
            HTTP_QATAPI_waring += 1
            if col_event[index] == '已开单':
                HTTP_QATAPI_Event += 1
            HTTP_QATAPI_object_list.append(objId)
        elif 'HTTP_ASystem_ERROR_CD_' in objId:
            HTTP_QAT_error_waring += 1
            if col_event[index] == '已开单':
                HTTP_QAT_error_Event += 1
            HTTP_QAT_error_object_list.append(objId)
        # ----------------------------------update 20190829----------------------------------
        elif 'QRC_' in objId and '_IM' in objId:
            QRC_APP_warning += 1
            if col_event[index] == '已开单':
                QRC_APP_Event += 1
            QRC_APP_list.append(objId)

        elif 'QAT_' in objId and '_B' in objId:
            QAT_Institution_warning += 1
            if col_event[index] == '已开单':
                QAT_Institution_Event += 1
            QAT_Institution_list.append(objId)

        elif 'QTT_' in objId and '_B' in objId:
            QTT_Institution_warning += 1
            if col_event[index] == '已开单':
                QTT_Institution_Event += 1
            QTT_Institution_list.append(objId)

        elif 'UPACP_' in objId and '_R' in objId:
            UPACP_Institution_warning += 1
            if col_event[index] == '已开单':
                UPACP_Institution_Event += 1
            UPACP_Institution_list.append(objId)

        elif 'UPACP_' in objId and '_A' in objId:
            UPACP_Institution_warning += 1
            if col_event[index] == '已开单':
                UPACP_Institution_Event += 1
            UPACP_Institution_list.append(objId)
        # ----------------------------------update 20190829----------------------------------
        # ----------------------------------update 20190829 新增客结机构130----------------------------------
        elif 'FS_T_' in objId and '_A' in objId:
            continue
        elif 'FS_T_' in objId and '_R' in objId:
            continue
        elif 'FS_A_' in objId and '_R' in objId:
            continue
        elif 'FS_A_' in objId and '_A' in objId:
            continue
        elif 'FS_' in objId and '_A' in objId:
            FS_Institution_warning += 1
            if col_event[index] == '已开单':
                FS_Institution_Event += 1
            FS_Institution_list.append(objId)
        elif 'FS_' in objId and '_R' in objId:
            FS_Institution_warning += 1
            if col_event[index] == '已开单':
                FS_Institution_Event += 1
            FS_Institution_list.append(objId)
        # ----------------------------------update 20190829 新增客结机构130----------------------------------
        # ----------------------------------update 20190917 新增UPACP交易中心----------------------------------
        elif 'UPACPTRANS_orderN_' in objId or 'UPACPNOTIFY_notice_' in objId:
            UPACP_Center_warning+=1
            if col_event[index] == '已开单':
                UPACP_Center_Event += 1
            UPACP_Center_list.append(objId)

        # ----------------------------------update 20190917 新增UPACP交易中心----------------------------------

        index += 1

    UPACP_All_object = len(set(UPACP_All_object_list))
    UPACP_All_object_str = ",".join(set(UPACP_All_object_list))
    UPACP_Type_object = len(set(UPACP_Type_object_list))
    UPACP_Type_object_str = ",".join(set(UPACP_All_object_list))
    UPACP_procsys_object = len(set(UPACP_procsys_object_list))
    UPACP_procsys_object_str = ",".join(set(UPACP_procsys_object_list))
    UPACP_client_id_object = len(set(UPACP_client_id_object_list))
    UPACP_client_id_object_str = ",".join(set(UPACP_client_id_object_list))
    UPACP_Merchant_object = len(set(UPACP_Merchant_object_list))
    UPACP_Merchant_object_str = ",".join(set(UPACP_Merchant_object_list))
    HTTP_UPACP_object = len(set(HTTP_UPACP_object_list))
    HTTP_UPACP_object_str = ",".join(set(HTTP_UPACP_object_list))
    HTTP_UPACP_qzhlw_object = len(set(HTTP_UPACP_qzhlw_object_list))
    HTTP_UPACP_qzhlw_object_str = ",".join(set(HTTP_UPACP_qzhlw_object_list))

    QRC_QRC_object = len(set(QRC_QRC_object_list))
    QRC_QRC_object_str = ",".join(set(QRC_QRC_object_list))
    QRC_Center_object = len(set(QRC_Center_object_list))
    QRC_Center_object_str = ",".join(set(QRC_Center_object_list))
    QRC_Type_object = len(set(QRC_Type_object_list))
    QRC_Type_object_str = ",".join(set(QRC_Type_object_list))
    QRC_Order_object = len(set(QRC_Order_object_list))
    QRC_Order_object_str = ",".join(set(QRC_Order_object_list))
    HTTP_QRC_All_object = len(set(HTTP_QRC_All_object_list))
    HTTP_QRC_All_object_str = ",".join(set(HTTP_QRC_All_object_list))
    HTTP_QRC_zxhlw_object = len(set(HTTP_QRC_zxhlw_object_list))
    HTTP_QRC_zxhlw_object_str = ",".join(set(HTTP_QRC_zxhlw_object_list))
    HTTP_QRC_error_object = len(set(HTTP_QRC_error_object_list))
    HTTP_QRC_error_object_str = ",".join(set(HTTP_QRC_error_object_list))

    FS_FS_object = len(set(FS_FS_object_list))
    FS_FS_object_str = ",".join(set(FS_FS_object_list))
    FS_Center_object = len(set(FS_Center_object_list))
    FS_Center_object_str = ",".join(set(FS_Center_object_list))
    FS_Type_object = len(set(FS_Type_object_list))
    FS_Type_object_str = ",".join(set(FS_Type_object_list))
    FS_procSys_object = len(set(FS_procSys_object_list))
    FS_procSys_object_str = ",".join(set(FS_procSys_object_list))
    HTTP_fsasAll_object = len(set(HTTP_fsasAll_object_list))
    HTTP_fsasAll_object_str = ",".join(set(HTTP_fsasAll_object_list))
    HTTP_FS_qzhlw_object = len(set(HTTP_FS_qzhlw_object_list))
    HTTP_FS_qzhlw_object_str = ",".join(set(FS_FS_object_list))
    HTTP_FS_error_object = len(set(HTTP_FS_error_object_list))
    HTTP_FS_error_object_str = ",".join(set(HTTP_FS_error_object_list))

    QTT_QTTB_object = len(set(QTT_QTTB_object_list))
    QTT_QTTB_object_str = ",".join(set(QTT_QTTB_object_list))
    QTT_Center_object = len(set(QTT_Center_object_list))
    QTT_Center_object_str = ",".join(set(QTT_Center_object_list))
    QTT_Type_object = len(set(QTT_Type_object_list))
    QTT_Type_object_str = ",".join(set(QTT_Type_object_list))
    HTTP_QTTall_object = len(set(HTTP_QTTall_object_list))
    HTTP_QTTall_object_str = ",".join(set(HTTP_QTTall_object_list))
    HTTP_QTTdomain_object = len(set(HTTP_QTTdomain_object_list))
    HTTP_QTTdomain_object_str = ",".join(set(HTTP_QTTdomain_object_list))
    HTTP_QTThlwqz_object = len(set(HTTP_QTThlwqz_object_list))
    HTTP_QTThlwqz_object_str = ",".join(set(HTTP_QTThlwqz_object_list))
    HTTP_QTThlw_qz_object = len(set(HTTP_QTThlw_qz_object_list))
    HTTP_QTThlw_qz_object_str = ",".join(set(HTTP_QTThlw_qz_object_list))
    HTTP_QTT_error_object = len(set(HTTP_QTT_error_object_list))
    HTTP_QTT_error_object_str = ",".join(set(HTTP_QTT_error_object_list))

    QAT_QAT_object = len(set(QAT_QAT_object_list))
    QAT_QAT_object_str = ",".join(set(QAT_QAT_object_list))
    QAT_Center_object = len(set(QAT_Center_object_list))
    QAT_Center_object_str = ",".join(set(QAT_Center_object_list))
    QAT_Type_object = len(set(QAT_Type_object_list))
    QAT_Type_object_str = ",".join(set(QAT_Type_object_list))
    HTTP_QATall_object = len(set(HTTP_QATall_object_list))
    HTTP_QATall_object_str = ",".join(set(HTTP_QATall_object_list))
    HTTP_QATdomain_object = len(set(HTTP_QATdomain_object_list))
    HTTP_QATdomain_object_str = ",".join(set(HTTP_QATdomain_object_list))
    HTTP_QATAPI_object = len(set(HTTP_QATAPI_object_list))
    HTTP_QATAPI_object_str = ",".join(set(HTTP_QATAPI_object_list))
    HTTP_QAT_error_object = len(set(HTTP_QAT_error_object_list))
    HTTP_QAT_error_object_str = ",".join(set(HTTP_QAT_error_object_list))

    # ----------------------------------update 20190829----------------------------------
    UPACP_Institution_object = len(set(UPACP_Institution_list))
    QRC_APP_object = len(set(QRC_APP_list))
    QTT_Institution_object = len(set(QTT_Institution_list))
    QAT_Institution_object = len(set(QAT_Institution_list))
    # ----------------------------------update 20190829----------------------------------
    # ----------------------------------update 20190902----------------------------------
    FS_Institution_object = len(set(FS_Institution_list))
    # ----------------------------------update 20190902----------------------------------
    # ----------------------------------update 20190917----------------------------------
    UPACP_Center_object = len(set(UPACP_Center_list))
    # ----------------------------------update 20190917----------------------------------

    # 写入excel特定位置
    fileModstr = str(input('输入模板excel名：'))
    fileMod = fileModstr
    nowTime = datetime.now().strftime("%m%d")
    wb = load_workbook(fileMod)
    ws = wb.worksheets[0]

    ws.cell(1,7).value = str(nowTime)+'告警数'
    # ws.write(0,6,str(nowTime)+'告警数')

    ws.cell(2, 7).value= UPACP_All_Warning
    ws.cell(2, 9).value= UPACP_All_Event
    ws.cell(2, 10).value= UPACP_All_object

    ws.cell(4, 7).value=  UPACP_Type_Warning
    ws.cell(4, 9).value=  UPACP_Type_Event
    ws.cell(4, 10).value= UPACP_Type_object

    ws.cell(5, 7).value= UPACP_procsys_Warning
    ws.cell(5, 9).value= UPACP_procsys_Event
    ws.cell(5, 10).value=UPACP_procsys_object

    ws.cell(6, 7).value=  UPACP_client_id_Warning
    ws.cell(6, 9).value=  UPACP_client_id_Event
    ws.cell(6, 10).value= UPACP_client_id_object

    ws.cell(7, 7).value=  UPACP_Merchant_Warning
    ws.cell(7, 9).value=  UPACP_Merchant_Event
    ws.cell(7, 10).value= UPACP_Merchant_object

    ws.cell(9, 7).value= HTTP_UPACP_Warning
    ws.cell(9, 9).value= HTTP_UPACP_Event
    ws.cell(9, 10).value=HTTP_UPACP_object

    ws.cell(11, 7).value= HTTP_UPACP_qzhlw_Warning
    ws.cell(11, 9).value= HTTP_UPACP_qzhlw_Event
    ws.cell(11, 10).value=HTTP_UPACP_qzhlw_object

    ws.cell(14, 7).value= QRC_QRC_Warning
    ws.cell(14, 9).value= QRC_QRC_Event
    ws.cell(14, 10).value=QRC_QRC_object

    ws.cell(15, 7).value= QRC_Center_Warning
    ws.cell(15, 9).value= QRC_Center_Event
    ws.cell(15, 10).value=QRC_Center_object

    ws.cell(16, 7).value= QRC_Type_Warning
    ws.cell(16, 9).value= QRC_Type_Event
    ws.cell(16, 10).value=QRC_Type_object

    ws.cell(17, 7).value= QRC_Order_Warning
    ws.cell(17, 9).value= QRC_Order_Event
    ws.cell(17, 10).value=QRC_Order_object

    ws.cell(21, 7).value= HTTP_QRC_All_Warning
    ws.cell(21, 9).value= HTTP_QRC_All_Event
    ws.cell(21, 10).value=HTTP_QRC_All_object

    ws.cell(23, 7).value= HTTP_QRC_zxhlw_Warning
    ws.cell(23, 9).value= HTTP_QRC_zxhlw_Event
    ws.cell(23, 10).value=HTTP_QRC_zxhlw_object

    ws.cell(25, 7).value= HTTP_QRC_error_Warning
    ws.cell(25, 9).value= HTTP_QRC_error_Event
    ws.cell(25, 10).value=HTTP_QRC_error_object

    ws.cell(26, 7).value= FS_FS_waring
    ws.cell(26, 9).value= FS_FS_Event
    ws.cell(26, 10).value=FS_FS_object

    ws.cell(27, 7).value= FS_Center_waring
    ws.cell(27, 9).value= FS_Center_Event
    ws.cell(27, 10).value=FS_Center_object

    ws.cell(28, 7).value= FS_Type_waring
    ws.cell(28, 9).value= FS_Type_Event
    ws.cell(28, 10).value=FS_Type_object

    ws.cell(29, 7).value= FS_procSys_waring
    ws.cell(29, 9).value= FS_procSys_Event
    ws.cell(29, 10).value=FS_procSys_object

    ws.cell(31, 7).value= HTTP_fsasAll_waring
    ws.cell(31, 9).value= HTTP_fsasAll_Event
    ws.cell(31, 10).value=HTTP_fsasAll_object

    ws.cell(33, 7).value= HTTP_FS_qzhlw_waring
    ws.cell(33, 9).value= HTTP_FS_qzhlw_Event
    ws.cell(33, 10).value=HTTP_FS_qzhlw_object

    ws.cell(35, 7).value= HTTP_FS_error_waring
    ws.cell(35, 9).value= HTTP_FS_error_Event
    ws.cell(35, 10).value=HTTP_FS_error_object

    ws.cell(36, 7).value= QTT_QTTB_waring
    ws.cell(36, 9).value= QTT_QTTB_Event
    ws.cell(36, 10).value=QTT_QTTB_object

    ws.cell(37, 7).value= QTT_Center_waring
    ws.cell(37, 9).value= QTT_Center_Event
    ws.cell(37, 10).value=QTT_Center_object

    ws.cell(38, 7).value= QTT_Type_waring
    ws.cell(38, 9).value= QTT_Type_Event
    ws.cell(38,10).value= QTT_Type_object

    ws.cell(40, 7).value= HTTP_QTTall_waring
    ws.cell(40, 9).value= HTTP_QTTall_Event
    ws.cell(40, 10).value=HTTP_QTTall_object

    ws.cell(41, 7).value= HTTP_QTTdomain_waring
    ws.cell(41, 9).value= HTTP_QTTdomain_Event
    ws.cell(41, 10).value=HTTP_QTTdomain_object

    ws.cell(42, 7).value= HTTP_QTThlwqz_waring
    ws.cell(42, 9).value= HTTP_QTThlwqz_Event
    ws.cell(42, 10).value=HTTP_QTThlwqz_object

    ws.cell(43, 7).value= HTTP_QTThlw_qz_waring
    ws.cell(43, 9).value= HTTP_QTThlw_qz_Event
    ws.cell(43, 10).value=HTTP_QTThlw_qz_object

    ws.cell(44, 7).value= HTTP_QTT_error_waring
    ws.cell(44, 9).value= HTTP_QTT_error_Event
    ws.cell(44, 10).value=HTTP_QTT_error_object

    ws.cell(45, 7).value= QAT_QAT_waring
    ws.cell(45, 9).value= QAT_QAT_Event
    ws.cell(45, 10).value=QAT_QAT_object

    ws.cell(46, 7).value= QAT_Center_waring
    ws.cell(46, 9).value= QAT_Center_Event
    ws.cell(46, 10).value=QAT_Center_object

    ws.cell(47, 7).value= QAT_Type_waring
    ws.cell(47, 9).value= QAT_Type_Event
    ws.cell(47, 10).value=QAT_Type_object

    ws.cell(49,7).value= HTTP_QATall_waring
    ws.cell(49, 9).value= HTTP_QATall_Event
    ws.cell(49, 10).value=HTTP_QATall_object

    ws.cell(51, 7).value= HTTP_QATdomain_waring
    ws.cell(51, 9).value= HTTP_QATdomain_Event
    ws.cell(51, 10).value=HTTP_QATdomain_object

    ws.cell(52, 7).value= HTTP_QATAPI_waring
    ws.cell(52, 9).value= HTTP_QATAPI_Event
    ws.cell(52, 10).value=HTTP_QATAPI_object

    ws.cell(53, 7).value= HTTP_QAT_error_waring
    ws.cell(53, 9).value= HTTP_QAT_error_Event
    ws.cell(53, 10).value=HTTP_QAT_error_object
    # ----------------------------------update 20190829----------------------------------
    ws.cell(8, 7).value = UPACP_Institution_warning
    ws.cell(8, 9).value = UPACP_Institution_Event
    ws.cell(8, 10).value = UPACP_Institution_object

    ws.cell(19, 7).value = QRC_APP_warning
    ws.cell(19, 9).value = QRC_APP_Event
    ws.cell(19, 10).value = QRC_APP_object

    ws.cell(39, 7).value = QTT_Institution_warning
    ws.cell(39, 9).value = QTT_Institution_Event
    ws.cell(39, 10).value = QTT_Institution_object

    ws.cell(48, 7).value = QAT_Institution_warning
    ws.cell(48, 9).value = QAT_Institution_Event
    ws.cell(48, 10).value = QAT_Institution_object

    # ----------------------------------update 20190902----------------------------------
    ws.cell(30, 7).value = FS_Institution_warning
    ws.cell(30, 9).value = FS_Institution_Event
    ws.cell(30, 10).value = FS_Institution_object

    # ----------------------------------update 20190917----------------------------------
    ws.cell(3, 7).value = UPACP_Center_warning
    ws.cell(3, 9).value = UPACP_Center_Event
    ws.cell(3, 10).value = UPACP_Center_object

    #-----------------------------------写入sheet2----------------------------------------
    ws1 = wb.worksheets[1]
    row = 2
    for i in set(UPACP_All_object_list):
        ws1.cell(row,1).value='全渠道'
        ws1.cell(row,2).value='UPACP'
        ws1.cell(row,3).value='整体对象监控'
        ws1.cell(row,4).value=i
        ws1.cell(row,5).value=str(UPACP_All_object_list.count(i))
        row += 1
    for i in set(UPACP_Type_object_list):
        ws1.cell(row,1).value= '全渠道'
        ws1.cell(row,2).value= 'UPACP'
        ws1.cell(row,3).value= '交易类型监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(UPACP_Type_object_list.count(i))
        row += 1
    for i in set(UPACP_procsys_object_list):
        ws1.cell(row,1).value= '全渠道'
        ws1.cell(row,2).value= 'UPACP'
        ws1.cell(row,3).value= '后端系统监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(UPACP_procsys_object_list.count(i))
        row += 1
    for i in set(UPACP_client_id_object_list):
        ws1.cell(row,1).value= '全渠道'
        ws1.cell(row,2).value= 'UPACP'
        ws1.cell(row,3).value= '网关监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(UPACP_client_id_object_list.count(i))
        row += 1
    for i in set(UPACP_Merchant_object_list):
        ws1.cell(row,1).value='全渠道'
        ws1.cell(row,2).value='UPACP'
        ws1.cell(row,3).value='商户监控'
        ws1.cell(row,4).value=i
        ws1.cell(row,5).value=str(UPACP_Merchant_object_list.count(i))
        row += 1
    for i in set(HTTP_UPACP_object_list):
        ws1.cell(row,1).value= '全渠道'
        ws1.cell(row,2).value= 'UPACP'
        ws1.cell(row,3).value= '整体对象监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_UPACP_object_list.count(i))
        row += 1
    for i in set(HTTP_UPACP_qzhlw_object_list):
        ws1.cell(row,1).value= '全渠道'
        ws1.cell(row,2).value= 'UPACP'
        ws1.cell(row,3).value= '接入域监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_UPACP_qzhlw_object_list.count(i))
        row += 1

    for i in set(QRC_QRC_object_list):
        ws1.cell(row,1).value= '二维码'
        ws1.cell(row,2).value= 'QRC'
        ws1.cell(row,3).value= '整体对象监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QRC_QRC_object_list.count(i))
        row += 1
    for i in set(QRC_Center_object_list):
        ws1.cell(row,1).value= '二维码'
        ws1.cell(row,2).value= 'QRC'
        ws1.cell(row,3).value= '交易中心监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QRC_Center_object_list.count(i))
        row += 1
    for i in set(QRC_Type_object_list):
        ws1.cell(row,1).value= '二维码'
        ws1.cell(row,2).value= 'QRC'
        ws1.cell(row,3).value= '交易类型监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QRC_Type_object_list.count(i))
        row += 1
    for i in set(QRC_Order_object_list):
        ws1.cell(row,1).value= '二维码'
        ws1.cell(row,2).value= 'QRC'
        ws1.cell(row,3).value= '订单类型监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QRC_Order_object_list.count(i))
        row += 1
    for i in set(HTTP_QRC_All_object_list):
        ws1.cell(row,1).value= '二维码'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '整体对象监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QRC_All_object_list.count(i))
        row += 1
    for i in set(HTTP_QRC_zxhlw_object_list):
        ws1.cell(row,1).value= '二维码'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '接入域监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QRC_zxhlw_object_list.count(i))
        row += 1
    for i in set(HTTP_QRC_error_object_list):
        ws1.cell(row,1).value='二维码'
        ws1.cell(row,2).value='HTTP'
        ws1.cell(row,3).value='错误码监控'
        ws1.cell(row,4).value=i
        ws1.cell(row,5).value=str(HTTP_QRC_error_object_list.count(i))
        row += 1

    for i in set(FS_FS_object_list):
        ws1.cell(row,1).value= '客结'
        ws1.cell(row,2).value= 'FS'
        ws1.cell(row,3).value= '整体对象监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(FS_FS_object_list.count(i))
        row += 1
    for i in set(FS_Center_object_list):
        ws1.cell(row,1).value= '客结'
        ws1.cell(row,2).value= 'FS'
        ws1.cell(row,3).value= '交易中心监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(FS_Center_object_list.count(i))
        row += 1
    for i in set(FS_Type_object_list):
        ws1.cell(row,1).value= '客结'
        ws1.cell(row,2).value= 'FS'
        ws1.cell(row,3).value= '交易类型监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(FS_Type_object_list.count(i))
        row += 1
    for i in set(FS_procSys_object_list):
        ws1.cell(row,1).value= '客结'
        ws1.cell(row,2).value= 'FS'
        ws1.cell(row,3).value= '后端系统监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(FS_procSys_object_list.count(i))
        row += 1
    for i in set(HTTP_fsasAll_object_list):
        ws1.cell(row,1).value= '客结'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '整体对象监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_fsasAll_object_list.count(i))
        row += 1
    for i in set(HTTP_FS_qzhlw_object_list):
        ws1.cell(row,1).value= '客结'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '接入域监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_FS_qzhlw_object_list.count(i))
        row += 1
    for i in set(HTTP_FS_error_object_list):
        ws1.cell(row,1).value= '客结'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '错误码监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_FS_error_object_list.count(i))
        row += 1

    for i in set(QTT_QTTB_object_list):
        ws1.cell(row,1).value= '条码T'
        ws1.cell(row,2).value= 'QTT'
        ws1.cell(row,3).value= '整体对象监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QTT_QTTB_object_list.count(i))
        row += 1
    for i in set(QTT_Center_object_list):
        ws1.cell(row,1).value= '条码T'
        ws1.cell(row,2).value= 'QTT'
        ws1.cell(row,3).value= '交易中心监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QTT_Center_object_list.count(i))
        row += 1
    for i in set(QTT_Type_object_list):
        ws1.cell(row,1).value= '条码T'
        ws1.cell(row,2).value= 'QTT'
        ws1.cell(row,3).value= '交易类型监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QTT_Type_object_list.count(i))
        row += 1
    for i in set(HTTP_QTTall_object_list):
        ws1.cell(row,1).value= '条码T'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '整体对象监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QTTall_object_list.count(i))
        row += 1
    for i in set(HTTP_QTTdomain_object_list):
        ws1.cell(row,1).value= '条码T'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '域名监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QTTdomain_object_list.count(i))
        row += 1
    for i in set(HTTP_QTThlwqz_object_list):
        ws1.cell(row,1).value= '条码T'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '域名监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QTThlwqz_object_list.count(i))
        row += 1
    for i in set(HTTP_QTThlw_qz_object_list):
        ws1.cell(row,1).value= '条码T'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= 'API监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QTThlw_qz_object_list.count(i))
        row += 1
    for i in set(HTTP_QTT_error_object_list):
        ws1.cell(row,1).value= '条码T'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '错误码监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QTT_error_object_list.count(i))
        row += 1

    for i in set(QAT_QAT_object_list):
        ws1.cell(row,1).value= '条码A'
        ws1.cell(row,2).value= 'QAT'
        ws1.cell(row,3).value= '整体对象监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QAT_QAT_object_list.count(i))
        row += 1
    for i in set(QAT_Center_object_list):
        ws1.cell(row,1).value= '条码A'
        ws1.cell(row,2).value= 'QAT'
        ws1.cell(row,3).value= '交易中心监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QAT_Center_object_list.count(i))
        row += 1
    for i in set(QAT_Type_object_list):
        ws1.cell(row,1).value= '条码A'
        ws1.cell(row,2).value= 'QAT'
        ws1.cell(row,3).value= '交易类型监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QAT_Type_object_list.count(i))
        row += 1
    for i in set(HTTP_QATall_object_list):
        ws1.cell(row,1).value= '条码A'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '整体对象监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QATall_object_list.count(i))
        row += 1
    for i in set(HTTP_QATdomain_object_list):
        ws1.cell(row,1).value= '条码A'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '接入域监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QATdomain_object_list.count(i))
        row += 1
    for i in set(HTTP_QATAPI_object_list):
        ws1.cell(row,1).value= '条码A'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= 'API监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QATAPI_object_list.count(i))
        row += 1
    for i in set(HTTP_QAT_error_object_list):
        ws1.cell(row,1).value= '条码A'
        ws1.cell(row,2).value= 'HTTP'
        ws1.cell(row,3).value= '错误码监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(HTTP_QAT_error_object_list.count(i))
        row += 1
    # ----------------------------------update 20190829----------------------------------
    for i in set(UPACP_Institution_list):
        ws1.cell(row,1).value= '全渠道'
        ws1.cell(row,2).value= 'UPACP'
        ws1.cell(row,3).value= '机构监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(UPACP_Institution_list.count(i))
        row += 1

    for i in set(QRC_APP_list):
        ws1.cell(row,1).value= '二维码'
        ws1.cell(row,2).value= 'QRC'
        ws1.cell(row,3).value= '收单机构监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QRC_APP_list.count(i))
        row += 1

    for i in set(QTT_Institution_list):
        ws1.cell(row,1).value= '条码T'
        ws1.cell(row,2).value= 'QTT'
        ws1.cell(row,3).value= '机构监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QTT_Institution_list.count(i))
        row += 1

    for i in set(QAT_Institution_list):
        ws1.cell(row,1).value= '条码A'
        ws1.cell(row,2).value= 'QAT'
        ws1.cell(row,3).value= '机构监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(QAT_Institution_list.count(i))
        row += 1
    # ----------------------------------update 20190829----------------------------------
    # ----------------------------------update 20190902----------------------------------
    for i in set(FS_Institution_list):
        ws1.cell(row,1).value= '客结'
        ws1.cell(row,2).value= 'FS'
        ws1.cell(row,3).value= '机构监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(FS_Institution_list.count(i))
        row += 1
    # ----------------------------------update 20190902----------------------------------
    # ----------------------------------update 20190917----------------------------------
    for i in set(UPACP_Center_list):
        ws1.cell(row,1).value= '全渠道'
        ws1.cell(row,2).value= 'UPACP'
        ws1.cell(row,3).value= '交易中心监控'
        ws1.cell(row,4).value= i
        ws1.cell(row,5).value= str(UPACP_Center_list.count(i))
        row += 1
    # ----------------------------------update 20190917----------------------------------

    fileName = "告警分析" + nowTime + '.xlsx'
    wb.save(fileName)


if __name__ == '__main__':
    InfoFile = str(input("输入信息excel表："))
    col_content = readExcel(InfoFile, 12)
    col_event = readExcel(InfoFile, 44)
    coll = extractData(col_content)
    writeExcelAnalyse(col_content, col_event)
    writeExcelRule(coll)
    # print(extractData(col))
    # print(len(extractData(col)))
