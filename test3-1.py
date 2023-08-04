from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

NowTime = datetime.now().strftime("%m%d")
if __name__ == '__main__':
    wb = load_workbook("C:\\Users\\yangzh\\Desktop\\up\\20190829\\告警分析-mod.xlsx")
    targetFile = "告警分析"+str(NowTime)+".xlsx"
    # cpy_sheet1 = wb.copy_worksheet(wb.worksheets[0])
    # cpy_sheet2 = wb.copy_worksheet(wb.worksheets[1])
    # cpy_sheet1.cell(5,1).value = 'nihao'
    # cpy_sheet2.cell(2,3).value = 'jaskdf'
    sheet1 = wb.worksheets[0]
    sheet2 = wb.worksheets[1]
    sheet1.cell(3,5).value='asffa'
    sheet2.cell(12,3).value = 'dasfas'
    wb.save(targetFile)
